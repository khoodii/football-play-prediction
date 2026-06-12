import json, re as _re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ILLEGAL=_re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def clean(v):
    return _ILLEGAL.sub("", v) if isinstance(v,str) else v

rows=json.load(open("/home/claude/parsed_v2.json"))

COLS=[("game","Opponent"),("drive","Drive"),("play_idx","Play #"),
      ("down","Down"),("dist_bucket","Dist (S/M/L)"),("dist_explicit","Dist (exact)"),
      ("goal_to_go","Goal"),("formation","Formation / Strength"),
      ("play_number","Play No."),("play_name","Play Call"),("play_family","Family"),
      ("broken_play","Broken"),("direction","Direction"),
      ("box_count","Box"),("shell_presnap","Shell (pre-snap)"),
      ("coverage_postsnap","Coverage (post-snap)"),("blitz","Blitz"),
      ("outcome","Result"),("yards","Yards"),("flag","Flag"),
      ("review","Review"),("user_note","Your note"),("raw","Raw entry")]

wb=Workbook()
ws=wb.active; ws.title="Plays"

HEAD=Font(name="Arial",bold=True,color="FFFFFF",size=10)
HFILL=PatternFill("solid",fgColor="1F3864")
BODY=Font(name="Arial",size=10)
REVIEW_FILL=PatternFill("solid",fgColor="FFF2CC")   # amber: needs a human look
EMPTY_FILL=PatternFill("solid",fgColor="F2F2F2")     # grey: empty row
NEEDS_FILL=PatternFill("solid",fgColor="F8CBAD")     # red: conflict, your 2nd pass
thin=Side(style="thin",color="D9D9D9")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

for c,(_,label) in enumerate(COLS,1):
    cell=ws.cell(1,c,label); cell.font=HEAD; cell.fill=HFILL
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=BORDER

for r,row in enumerate(rows,2):
    rv=row.get("review")
    for c,(key,_) in enumerate(COLS,1):
        v=row.get(key)
        v="" if v is None else clean(v)   # TRUE blanks, never 0/placeholder
        cell=ws.cell(r,c,v); cell.font=BODY; cell.border=BORDER
        cell.alignment=Alignment(horizontal="left",vertical="center",
                                 wrap_text=(key=="raw"))
        if rv=="empty": cell.fill=EMPTY_FILL
        elif rv=="needs_user": cell.fill=NEEDS_FILL
        elif rv: cell.fill=REVIEW_FILL

widths=[12,6,6,6,11,11,6,20,8,12,9,8,9,6,16,18,16,12,7,18,16,40,46]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="D2"
ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
ws.row_dimensions[1].height=30

# ---------- Schema sheet ----------
sc=wb.create_sheet("Schema")
sc["A1"]="Play-Log Schema & Normalization Rules"; sc["A1"].font=Font(name="Arial",bold=True,size=13)
schema=[
 ("Column","Meaning / normalization"),
 ("Opponent","Source game file. Kearsley = play-tracking only (not a 2026 opponent)."),
 ("Drive","Sequential drive within the game, from 'drive N' headers."),
 ("Play #","Sequence of the play within the game."),
 ("Down","1-4. Taken from leading 1/2/3/4 or '1st/2nd/...'."),
 ("Dist (S/M/L)","Short / Medium / Long bucket when written as /s/ /m/ /l/."),
 ("Dist (exact)","Exact yards when written explicitly ('1 and 15' -> 15)."),
 ("Goal","Y when 'goal' / goal-line situation noted."),
 ("Formation / Strength","Offensive tags: rip, louie, left/right strong-weak, plus/minus, trip, empty, etc."),
 ("Play No.","Offensive play number (47, 23, 61...). Number families map to concepts via your playbook (a later step)."),
 ("Play Call","Offensive concept (power, dragon, fade...) or generic type (run / pass / qb run) when no concept was named. THIS IS THE PREDICTION TARGET."),
 ("Family","run / pass / qb_run / qb_pass — only set when textually unambiguous. Blank for named numbers pending playbook mapping (not guessed)."),
 ("Broken","Y when the called play broke down ('broken play')."),
 ("Direction","left / right / middle."),
 ("Box","Defenders in box, 5-8. CARRIES FORWARD until you write a new value (sticky)."),
 ("Shell (pre-snap)","Safety picture shown BEFORE the snap: 2-high (deuce), 1-high, 3-high, 4-deep."),
 ("Coverage (post-snap)","Coverage actually PLAYED: cover 0-4, man, zone. solid = cover 3; '1 high' = cover 3; deuce = 2-high shell that plays off solid."),
 ("Blitz","Pressure noted: a gap, double A, sim pressure, mike blitz, etc."),
 ("Result","Standardized outcome: TD, INT, sack, fumble, incomplete, complete, no_gain, first_down."),
 ("Yards","Yards gained/lost when a number was recorded."),
 ("Flag","penalty:* / dead_call / play_missed / review — non-play or noted events."),
 ("Review","Blank = clean. needs_user = your note conflicts with the entry, needs your 2nd look. result_only = real play, only result logged. presnap_or_note = alignment/observation. no_down = play w/o down. empty = blank. observation_only = note."),
 ("Your note","Your correction / clarification folded in from the manual pass."),
 ("Raw entry","Your original line, untouched, so every row is auditable."),
]
for r,(a,b) in enumerate(schema,3):
    ca=sc.cell(r,1,a); cb=sc.cell(r,2,b)
    ca.font=Font(name="Arial",bold=(r==3),size=10); cb.font=Font(name="Arial",bold=(r==3),size=10)
    ca.alignment=Alignment(vertical="top"); cb.alignment=Alignment(vertical="top",wrap_text=True)
    if r==3:
        ca.fill=HFILL; cb.fill=HFILL; ca.font=HEAD; cb.font=HEAD
sc.column_dimensions["A"].width=22; sc.column_dimensions["B"].width=95

# ---------- Summary sheet (live formulas) ----------
sm=wb.create_sheet("Summary"); sm["A1"]="Per-Game Parse Summary"
sm["A1"].font=Font(name="Arial",bold=True,size=13)
hdr=["Opponent","Total rows","Clean plays","Play call present","Needs review"]
for c,h in enumerate(hdr,1):
    cell=sm.cell(3,c,h); cell.font=HEAD; cell.fill=HFILL
    cell.alignment=Alignment(horizontal="center",wrap_text=True)
games=sorted({r["game"] for r in rows})
last=len(rows)+1
for i,g in enumerate(games,4):
    sm.cell(i,1,g).font=BODY
    sm.cell(i,2,f'=COUNTIF(Plays!A2:A{last},A{i})').font=BODY
    sm.cell(i,3,f'=COUNTIFS(Plays!A2:A{last},A{i},Plays!U2:U{last},"")').font=BODY
    sm.cell(i,4,f'=COUNTIFS(Plays!A2:A{last},A{i},Plays!J2:J{last},"<>")').font=BODY
    sm.cell(i,5,f'=COUNTIFS(Plays!A2:A{last},A{i},Plays!U2:U{last},"<>")').font=BODY
tot=4+len(games)
sm.cell(tot,1,"TOTAL").font=Font(name="Arial",bold=True,size=10)
for c in range(2,6):
    L=get_column_letter(c)
    sm.cell(tot,c,f'=SUM({L}4:{L}{tot-1})').font=Font(name="Arial",bold=True,size=10)
for c,w in enumerate([14,11,12,17,13],1): sm.column_dimensions[get_column_letter(c)].width=w

wb.save("/home/claude/play_data_v2.xlsx")
print("saved play_data_v2.xlsx with", len(rows), "rows")
